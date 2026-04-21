from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
DASHBOARD_TEMPLATE_PATH = ROOT / "Eluru_Village_Issues_Dashboard.xlsx"
SOURCE_WORKBOOK_PATH = ROOT / "1.ELR_DIST-ABC_VILLAGES_CATEGORY_WISE_UPDATE_AS_ON_31.03.2026.xlsx"

ISSUE_TYPES = [
    "Caste Conflicts",
    "Political Issues",
    "Communal Issues",
    "General Issues",
]

ISSUE_SHEET_MAP = {
    "Cast Conflicts": "Caste Conflicts",
    "Political": "Political Issues",
    "Communal": "Communal Issues",
    "General": "General Issues",
}

ISSUE_LABEL_TO_TEMPLATE_COLUMN = {
    "Caste Conflicts": "E",
    "Political Issues": "F",
    "Communal Issues": "G",
    "General Issues": "H",
}

SUBDIVISION_ORDER = ["Eluru", "Nuzvidu", "J.R Gudem", "Polavaram"]
SUBDIVISION_DISPLAY_LABELS = {
    "Eluru": "Eluru Sub-Division",
    "Nuzvidu": "Nuzvidu Sub-Division",
    "J.R Gudem": "J.R. Gudem Sub-Div.",
    "Polavaram": "Polavaram Sub-Div.",
}

CATEGORY_ORDER = ["A", "B", "C"]
CATEGORY_ALERT_LABELS = {
    "A": "HIGH",
    "B": "WATCH",
    "C": "PEACEFUL",
}
CATEGORY_LABELS = {
    "A": "Category A",
    "B": "Category B",
    "C": "Category C",
}

PRESET_PARTY_COMBINATIONS = [
    "TDP vs YSRCP",
    "Internal YSRCP",
    "TDP vs JSP",
    "YSRCP vs JSP",
    "YSRCP vs Congress",
    "Internal TDP",
    "Social Media Abuse",
    "TDP vs Congress",
    "Internal JSP",
]

TOP_PS_LIMIT = 10
POLITICAL_BREAKDOWN_LIMIT = 10
POLITICAL_CHART_LIMIT = 7
ACTION_TRACKER_LIMIT = 15

CHART_COLORS = {
    "Caste Conflicts": "#4F81BD",
    "Political Issues": "#C0504D",
    "Communal Issues": "#9BBB59",
    "General Issues": "#8064A2",
    "Category A": "#C0504D",
    "Category B": "#F2A65A",
    "Category C": "#9BBB59",
}

CATEGORY_SEVERITY_RANK = {"A": 0, "B": 1, "C": 2}
ISSUE_TYPE_PRIORITY = {
    "Political Issues": 0,
    "Caste Conflicts": 1,
    "Communal Issues": 2,
    "General Issues": 3,
}

PARTY_COMBINATION_ALIASES = {
    "tdpvsysrcp": "TDP vs YSRCP",
    "tdpvsycp": "TDP vs YSRCP",
    "ysrcpvstdp": "TDP vs YSRCP",
    "ycpvstdp": "TDP vs YSRCP",
    "internalysrcp": "Internal YSRCP",
    "ysrcpinternal": "Internal YSRCP",
    "ycpinternal": "Internal YSRCP",
    "twogroupsinysrcp": "Internal YSRCP",
    "tdpvsjsp": "TDP vs JSP",
    "tdpvsjanasena": "TDP vs JSP",
    "janasenavstdp": "TDP vs JSP",
    "ysrcpvsjsp": "YSRCP vs JSP",
    "ycpvsjsp": "YSRCP vs JSP",
    "ysrcpvsjanasena": "YSRCP vs JSP",
    "ysrcpvscongress": "YSRCP vs Congress",
    "ycpvscongress": "YSRCP vs Congress",
    "internaltdp": "Internal TDP",
    "tdpinternal": "Internal TDP",
    "socialmediaabuse": "Social Media Abuse",
    "tdpvscongress": "TDP vs Congress",
    "internaljsp": "Internal JSP",
}

SAMPLE_PARTY_OVERRIDES = {
    ("Chebrole", "Chinna Vellamilli"): "TDP vs JSP",
    ("Nidamarru", "Krovidi"): "TDP vs JSP",
    ("Pedavegi", "Munduru village"): "TDP vs Congress",
    ("Velerupadu", "Velairpad village"): "YSRCP vs JSP",
    ("D.Tirumala", "Thimmapuram"): "Internal TDP",
    ("Nidamarru", "Gunaparru"): "Internal JSP",
}

_STATION_CATALOG: list[dict[str, Any]] | None = None
_STATION_LOOKUP: dict[str, str] | None = None
_STATION_SUBDIVISION_LOOKUP: dict[str, str] | None = None
_DEFAULT_REPORT_DATE: date | None = None
_SAMPLE_RECORDS_CACHE: list[dict[str, Any]] | None = None


def keyify(value: Any) -> str:
    return "".join(ch.lower() for ch in str(value or "") if ch.isalnum())


def collapse_whitespace(value: Any) -> str:
    text = str(value or "")
    return re.sub(r"\s+", " ", text).strip()


def shorten_text(value: Any, limit: int = 88) -> str:
    text = collapse_whitespace(value)
    if len(text) <= limit:
        return text

    clipped = text[:limit].rsplit(" ", 1)[0].strip()
    return f"{clipped or text[:limit].strip()}..."


def get_default_report_date() -> date:
    global _DEFAULT_REPORT_DATE

    if _DEFAULT_REPORT_DATE is not None:
        return _DEFAULT_REPORT_DATE

    workbook = load_workbook(DASHBOARD_TEMPLATE_PATH, data_only=True, read_only=True)
    title = workbook["Executive Dashboard"]["B2"].value
    workbook.close()

    match = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", str(title or ""))
    if match:
        day, month, year = (int(piece) for piece in match.groups())
        _DEFAULT_REPORT_DATE = date(year, month, day)
    else:
        _DEFAULT_REPORT_DATE = date.today()

    return _DEFAULT_REPORT_DATE


def parse_report_date(value: Any) -> date:
    if isinstance(value, date):
        return value

    text = collapse_whitespace(value)
    if not text:
        return get_default_report_date()

    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    return get_default_report_date()


def format_report_date(report_date: Any) -> str:
    parsed = parse_report_date(report_date)
    return parsed.strftime("%d.%m.%Y")


def ensure_station_catalog() -> None:
    global _STATION_CATALOG, _STATION_LOOKUP, _STATION_SUBDIVISION_LOOKUP

    if _STATION_CATALOG is not None and _STATION_LOOKUP is not None and _STATION_SUBDIVISION_LOOKUP is not None:
        return

    workbook = load_workbook(DASHBOARD_TEMPLATE_PATH, data_only=True, read_only=True)
    worksheet = workbook["PS-Wise Analysis"]

    catalog: list[dict[str, Any]] = []
    lookup: dict[str, str] = {}
    subdivision_lookup: dict[str, str] = {}

    for row in range(6, 40):
        serial = worksheet[f"B{row}"].value
        station_name = collapse_whitespace(worksheet[f"C{row}"].value)
        subdivision = collapse_whitespace(worksheet[f"D{row}"].value)
        if not serial or not station_name:
            continue

        catalog.append(
            {
                "serial": int(serial),
                "name": station_name,
                "subDivision": subdivision,
            }
        )
        lookup[keyify(station_name)] = station_name
        subdivision_lookup[station_name] = subdivision

    workbook.close()

    manual_aliases = {
        "eluruitn": "Eluru II Town",
        "eluruiitn": "Eluru II Town",
        "eluruiitown": "Eluru II Town",
        "eluruiitownps": "Eluru II Town",
        "eluruitown": "Eluru I Town",
        "eluruiitownpolicestation": "Eluru II Town",
        "dwarakatirumalaps": "D.Tirumala",
        "dwarakatirumala": "D.Tirumala",
        "dtirumala": "D.Tirumala",
        "kaikalurutownps": "Kaikaluru Town",
        "kaikalurutown": "Kaikaluru Town",
        "kaikalurururalps": "Kaikaluru Rural",
        "nidmarrups": "Nidamarru",
        "nidamarrups": "Nidamarru",
        "nuzvidruralps": "Nuzvid Rural",
        "nuzvidurural": "Nuzvid Rural",
        "nuzvidururalps": "Nuzvid Rural",
        "dendulurups": "Denduluru",
        "agiripallips": "Agiripalli",
        "mandavallips": "Mandavalli",
        "pedavegips": "Pedavegi",
        "dharmajigudemps": "Darmajigudem",
        "dharmajigudemps": "Darmajigudem",
        "dharmajigudem": "Darmajigudem",
        "velairpad": "Velerupadu",
        "velairpadvillage": "Velerupadu",
        "velerupadu": "Velerupadu",
        "jrgudem": "J.R Gudem",
        "tnarasapuram": "T.Narasapuram",
        "tnarasapuramm": "T.Narasapuram",
        "tnarasapuramps": "T.Narasapuram",
        "kukkunoor": "Kuknoor",
        "kukunoor": "Kuknoor",
        "kukkunoor": "Kuknoor",
        "eluruitown": "Eluru I Town",
        "eluruiiitown": "Eluru III Town",
    }

    for alias, canonical in manual_aliases.items():
        lookup[alias] = canonical

    _STATION_CATALOG = catalog
    _STATION_LOOKUP = lookup
    _STATION_SUBDIVISION_LOOKUP = subdivision_lookup


def get_station_catalog() -> list[dict[str, Any]]:
    ensure_station_catalog()
    return list(_STATION_CATALOG or [])


def get_station_lookup() -> dict[str, str]:
    ensure_station_catalog()
    return dict(_STATION_LOOKUP or {})


def get_station_subdivision_lookup() -> dict[str, str]:
    ensure_station_catalog()
    return dict(_STATION_SUBDIVISION_LOOKUP or {})


def normalize_police_station(value: Any) -> str:
    station = collapse_whitespace(value)
    if not station:
        return ""

    station_lookup = get_station_lookup()
    lookup_key = keyify(station)
    canonical = station_lookup.get(lookup_key)
    if canonical:
        return canonical

    for key, name in station_lookup.items():
        if lookup_key in key or key in lookup_key:
            return name

    return station


def infer_subdivision(police_station: Any) -> str:
    station = normalize_police_station(police_station)
    if not station:
        return ""

    subdivision_lookup = get_station_subdivision_lookup()
    return subdivision_lookup.get(station, "")


def normalize_issue_type(value: Any) -> str:
    text = collapse_whitespace(value)
    if not text:
        return ""

    if text in ISSUE_TYPES:
        return text

    mapped = ISSUE_SHEET_MAP.get(text)
    if mapped:
        return mapped

    lowered = text.lower()
    if "caste" in lowered:
        return "Caste Conflicts"
    if "politic" in lowered:
        return "Political Issues"
    if "communal" in lowered:
        return "Communal Issues"
    if "general" in lowered:
        return "General Issues"

    return text


def normalize_category(value: Any) -> str:
    text = collapse_whitespace(value).upper()
    if text in CATEGORY_ORDER:
        return text
    return ""


def normalize_party_combination(value: Any) -> str:
    text = collapse_whitespace(value)
    if not text:
        return ""

    alias = PARTY_COMBINATION_ALIASES.get(keyify(text))
    if alias:
        return alias

    return text


def detect_parties(text: str) -> dict[str, bool]:
    lowered = text.lower()
    return {
        "ysrcp": bool(re.search(r"\b(ysrcp|ycp)\b", lowered)),
        "tdp": "tdp" in lowered or "telugu desam" in lowered,
        "jsp": bool(re.search(r"\b(jsp|jana sena|janasena)\b", lowered)),
        "congress": "congress" in lowered,
    }


def derive_party_combination(record: dict[str, Any]) -> str:
    station = collapse_whitespace(record.get("policeStation"))
    village = collapse_whitespace(record.get("village"))
    override = SAMPLE_PARTY_OVERRIDES.get((station, village))
    if override:
        return override

    source_text = collapse_whitespace(
        " ".join(
            [
                record.get("issueSummary", ""),
                record.get("issueDetails", ""),
                record.get("remarks", ""),
                record.get("presentStatus", ""),
            ]
        )
    ).lower()

    parties = detect_parties(source_text)
    has_ysrcp = parties["ysrcp"]
    has_tdp = parties["tdp"]
    has_jsp = parties["jsp"]
    has_congress = parties["congress"]

    if "social media" in source_text:
        return "Social Media Abuse"

    if (
        "internal differences among ysrcp" in source_text
        or "two groups in ysrcp" in source_text
        or "two groups in ycp" in source_text
        or "both are ysrcp" in source_text
    ):
        return "Internal YSRCP"

    if (
        "two groups belonging to the telugu desam party" in source_text
        or "internal differences among tdp" in source_text
        or "two groups in tdp" in source_text
    ):
        return "Internal TDP"

    if "internal political conflict" in source_text and has_jsp and not has_tdp and not has_ysrcp and not has_congress:
        return "Internal JSP"

    if has_tdp and has_ysrcp and not has_jsp and not has_congress:
        return "TDP vs YSRCP"
    if has_tdp and has_jsp and not has_ysrcp and not has_congress:
        return "TDP vs JSP"
    if has_ysrcp and has_jsp and not has_tdp and not has_congress:
        return "YSRCP vs JSP"
    if has_ysrcp and has_congress and not has_tdp:
        return "YSRCP vs Congress"
    if has_tdp and has_congress and not has_ysrcp:
        return "TDP vs Congress"

    if has_ysrcp and not has_tdp and not has_jsp and not has_congress:
        return "Internal YSRCP"
    if has_tdp and not has_ysrcp and not has_jsp and not has_congress:
        return "Internal TDP"
    if has_jsp and not has_tdp and not has_ysrcp and not has_congress:
        return "Internal JSP"

    return ""


def auto_issue_summary(record: dict[str, Any]) -> str:
    supplied_summary = collapse_whitespace(record.get("issueSummary"))
    if supplied_summary:
        return shorten_text(supplied_summary, 88)

    issue_type = record.get("issueType", "")
    party_combination = collapse_whitespace(record.get("partyCombination"))
    village = collapse_whitespace(record.get("village"))
    details = collapse_whitespace(record.get("issueDetails"))

    if issue_type == "Political Issues" and party_combination:
        return shorten_text(f"{party_combination} issue - {village or 'Village'}", 88)

    if details:
        first_chunk = re.split(r"[.:;]", details, maxsplit=1)[0]
        cleaned = collapse_whitespace(first_chunk)
        if cleaned:
            return shorten_text(cleaned, 88)

    if village:
        return shorten_text(f"{issue_type or 'Village issue'} - {village}", 88)

    return shorten_text(issue_type or "Issue", 88)


def sanitize_record(raw_record: dict[str, Any], index: int = 0) -> dict[str, Any]:
    issue_type = normalize_issue_type(
        raw_record.get("issueType")
        or raw_record.get("issue_type")
        or raw_record.get("sheetName")
        or raw_record.get("sheet_name")
    )
    police_station = normalize_police_station(
        raw_record.get("policeStation")
        or raw_record.get("police_station")
        or raw_record.get("ps")
        or raw_record.get("station")
    )
    category = normalize_category(raw_record.get("category"))
    village = collapse_whitespace(raw_record.get("village") or raw_record.get("ward") or raw_record.get("location"))
    issue_details = collapse_whitespace(raw_record.get("issueDetails") or raw_record.get("issue") or raw_record.get("issue_detail"))
    remarks = collapse_whitespace(raw_record.get("remarks") or raw_record.get("remarksActionTaken") or raw_record.get("remark"))
    action_taken = collapse_whitespace(raw_record.get("actionTaken") or raw_record.get("action_taken") or remarks)
    present_status = collapse_whitespace(raw_record.get("presentStatus") or raw_record.get("present_status"))
    sub_division = collapse_whitespace(raw_record.get("subDivision") or raw_record.get("sub_division")) or infer_subdivision(police_station)
    issue_summary = collapse_whitespace(raw_record.get("issueSummary") or raw_record.get("issue_summary"))

    record = {
        "id": str(raw_record.get("id") or raw_record.get("recordId") or f"issue-{index + 1}"),
        "issueType": issue_type,
        "policeStation": police_station,
        "subDivision": sub_division,
        "village": village,
        "issueDetails": issue_details,
        "remarks": remarks,
        "category": category,
        "actionTaken": action_taken,
        "presentStatus": present_status,
        "issueSummary": issue_summary,
        "partyCombination": normalize_party_combination(
            raw_record.get("partyCombination") or raw_record.get("party_combination")
        ),
    }

    if issue_type == "Political Issues" and not record["partyCombination"]:
        record["partyCombination"] = derive_party_combination(record)

    record["issueSummary"] = auto_issue_summary(record)
    record["alertLevel"] = CATEGORY_ALERT_LABELS.get(category, "MONITOR")

    return record


def extract_sample_records() -> list[dict[str, Any]]:
    workbook = load_workbook(SOURCE_WORKBOOK_PATH, data_only=True, read_only=True)
    records: list[dict[str, Any]] = []

    for sheet_name, issue_type in ISSUE_SHEET_MAP.items():
        worksheet = workbook[sheet_name]
        for row in range(4, worksheet.max_row + 1):
            police_station = worksheet[f"B{row}"].value
            issue_text = worksheet[f"D{row}"].value
            if police_station is None and issue_text is None:
                continue

            record = sanitize_record(
                {
                    "id": f"{keyify(issue_type)}-{row}",
                    "issueType": issue_type,
                    "policeStation": police_station,
                    "village": worksheet[f"C{row}"].value,
                    "issueDetails": issue_text,
                    "remarks": worksheet[f"E{row}"].value,
                    "category": worksheet[f"F{row}"].value,
                    "actionTaken": worksheet[f"G{row}"].value,
                    "presentStatus": worksheet[f"H{row}"].value,
                },
                index=len(records),
            )

            if not record["issueType"] or not record["policeStation"]:
                continue
            records.append(record)

    workbook.close()
    return records


def load_sample_records() -> list[dict[str, Any]]:
    global _SAMPLE_RECORDS_CACHE

    if _SAMPLE_RECORDS_CACHE is None:
        _SAMPLE_RECORDS_CACHE = extract_sample_records()

    return [dict(record) for record in _SAMPLE_RECORDS_CACHE]


def build_bootstrap_payload() -> dict[str, Any]:
    return {
        "defaultReportDate": get_default_report_date().isoformat(),
        "defaultReportDateLabel": format_report_date(get_default_report_date()),
        "issueTypes": list(ISSUE_TYPES),
        "categories": [
            {"value": category, "label": CATEGORY_LABELS[category], "alert": CATEGORY_ALERT_LABELS[category]}
            for category in CATEGORY_ORDER
        ],
        "partyCombinations": list(PRESET_PARTY_COMBINATIONS),
        "policeStations": get_station_catalog(),
        "sampleRecords": load_sample_records(),
        "templateWorkbookName": DASHBOARD_TEMPLATE_PATH.name,
        "sourceWorkbookName": SOURCE_WORKBOOK_PATH.name,
    }


def severity_label(total: int) -> str:
    if total >= 8:
        return "HIGH"
    if total >= 4:
        return "MODERATE"
    if total >= 1:
        return "LOW"
    return "NIL"


def percentage_string(value: int, total: int, decimals: int = 1) -> str:
    if total <= 0:
        return f"{0:.{decimals}f}%"
    percentage = (value / total) * 100
    if decimals == 0:
        return f"{round(percentage):.0f}%"
    return f"{percentage:.{decimals}f}%"


def executive_card_insight(
    count: int,
    total: int,
    highest_value: int,
    *,
    prefer_highest: bool = True,
    tie_value: int | None = None,
) -> str:
    if count <= 0 or total <= 0:
        return "No issues reported"

    share = percentage_string(count, total, 1)
    if prefer_highest and count == highest_value and highest_value > 0:
        if tie_value is not None and tie_value == highest_value:
            return f"Tied highest ({share})"
        return f"Highest category ({share})"
    return f"{share} of total"


def political_status_label(categories: set[str], issue_count: int, statuses: list[str]) -> str:
    combined = " ".join(statuses).lower()

    if any(word in combined for word in ["high tension", "active rivalry", "serious tension"]):
        return "High tension areas"
    if any(word in combined for word in ["compromise", "compramised", "compromised"]):
        return "Compromised"
    if any(word in combined for word in ["peaceful", "no issue", "no issues"]):
        if categories == {"C"}:
            return "Peaceful"

    if "A" in categories:
        return "High tension areas" if issue_count >= 5 else "Active rivalry"
    if "B" in categories:
        return "Under watch" if issue_count >= 3 else "Moderate watch"
    if "C" in categories:
        return "Peaceful"
    return "Monitoring"


def make_blank_ps_row(station: dict[str, Any]) -> dict[str, Any]:
    return {
        "serial": station["serial"],
        "policeStation": station["name"],
        "subDivision": station["subDivision"],
        "casteConflicts": 0,
        "politicalIssues": 0,
        "communalIssues": 0,
        "generalIssues": 0,
        "totalIssues": 0,
        "severity": "NIL",
    }


def sanitize_records(raw_records: list[dict[str, Any]] | None) -> list[dict[str, Any]]:
    sanitized: list[dict[str, Any]] = []
    for index, record in enumerate(raw_records or []):
        item = sanitize_record(record or {}, index=index)
        if item["issueType"] and item["policeStation"]:
            sanitized.append(item)
    return sanitized


def build_dashboard_payload(raw_records: list[dict[str, Any]] | None, report_date: Any = None) -> dict[str, Any]:
    records = [sanitize_record(record or {}, index=index) for index, record in enumerate(raw_records or [])]
    records = [record for record in records if record["issueType"] and record["policeStation"]]

    report_date_obj = parse_report_date(report_date)
    report_date_label = format_report_date(report_date_obj)

    issue_totals = Counter({issue_type: 0 for issue_type in ISSUE_TYPES})
    subdivision_totals: dict[str, Counter[str]] = {
        subdivision: Counter({issue_type: 0 for issue_type in ISSUE_TYPES})
        for subdivision in SUBDIVISION_ORDER
    }
    ps_totals: dict[str, Counter[str]] = {
        station["name"]: Counter({issue_type: 0 for issue_type in ISSUE_TYPES})
        for station in get_station_catalog()
    }
    category_matrix: dict[str, Counter[str]] = {
        issue_type: Counter({category: 0 for category in CATEGORY_ORDER})
        for issue_type in ISSUE_TYPES
    }

    political_records: list[dict[str, Any]] = []

    for record in records:
        issue_type = record["issueType"]
        issue_totals[issue_type] += 1

        sub_division = record["subDivision"] or infer_subdivision(record["policeStation"])
        if sub_division in subdivision_totals:
            subdivision_totals[sub_division][issue_type] += 1

        if record["policeStation"] in ps_totals:
            ps_totals[record["policeStation"]][issue_type] += 1

        if record["category"] in CATEGORY_ORDER:
            category_matrix[issue_type][record["category"]] += 1

        if issue_type == "Political Issues":
            political_records.append(record)

    total_issues = sum(issue_totals.values())
    political_total = issue_totals["Political Issues"]
    general_total = issue_totals["General Issues"]
    caste_total = issue_totals["Caste Conflicts"]
    communal_total = issue_totals["Communal Issues"]
    caste_communal_total = caste_total + communal_total
    top_card_highest = max(political_total, general_total, caste_communal_total)

    subdivision_rows: list[dict[str, Any]] = []
    for subdivision in SUBDIVISION_ORDER:
        counts = subdivision_totals[subdivision]
        total = sum(counts.values())
        subdivision_rows.append(
            {
                "name": SUBDIVISION_DISPLAY_LABELS[subdivision],
                "shortName": subdivision,
                "casteConflicts": counts["Caste Conflicts"],
                "politicalIssues": counts["Political Issues"],
                "communalIssues": counts["Communal Issues"],
                "generalIssues": counts["General Issues"],
                "totalIssues": total,
                "share": percentage_string(total, total_issues, 1),
            }
        )

    executive_distribution = [
        {
            "label": issue_type,
            "value": issue_totals[issue_type],
            "color": CHART_COLORS[issue_type],
        }
        for issue_type in ISSUE_TYPES
    ]

    executive_chart_series = [
        {
            "label": issue_type,
            "color": CHART_COLORS[issue_type],
            "values": [
                subdivision_totals[subdivision][issue_type]
                for subdivision in SUBDIVISION_ORDER
            ],
        }
        for issue_type in ISSUE_TYPES
    ]

    ps_rows: list[dict[str, Any]] = []
    for station in get_station_catalog():
        counts = ps_totals[station["name"]]
        total = sum(counts.values())
        ps_rows.append(
            {
                "serial": station["serial"],
                "policeStation": station["name"],
                "subDivision": station["subDivision"],
                "casteConflicts": counts["Caste Conflicts"],
                "politicalIssues": counts["Political Issues"],
                "communalIssues": counts["Communal Issues"],
                "generalIssues": counts["General Issues"],
                "totalIssues": total,
                "severity": severity_label(total),
            }
        )

    top_stations = [
        {
            "policeStation": row["policeStation"],
            "totalIssues": row["totalIssues"],
        }
        for row in sorted(
            [row for row in ps_rows if row["totalIssues"] > 0],
            key=lambda row: (-row["totalIssues"], row["policeStation"]),
        )[:TOP_PS_LIMIT]
    ]

    category_rows: list[dict[str, Any]] = []
    category_totals = Counter({category: 0 for category in CATEGORY_ORDER})
    for issue_type in ISSUE_TYPES:
        counts = category_matrix[issue_type]
        total = sum(counts.values())
        for category in CATEGORY_ORDER:
            category_totals[category] += counts[category]

        category_rows.append(
            {
                "issueType": issue_type,
                "categoryA": counts["A"],
                "categoryB": counts["B"],
                "categoryC": counts["C"],
                "total": total,
                "percent": percentage_string(total, total_issues, 1),
            }
        )

    overall_category_total = sum(category_totals.values())
    category_split = [
        {
            "label": CATEGORY_LABELS[category],
            "value": category_totals[category],
            "color": CHART_COLORS[CATEGORY_LABELS[category]],
        }
        for category in CATEGORY_ORDER
    ]

    category_chart_series = [
        {
            "label": CATEGORY_LABELS["A"],
            "color": CHART_COLORS["Category A"],
            "values": [category_matrix[issue_type]["A"] for issue_type in ISSUE_TYPES],
        },
        {
            "label": CATEGORY_LABELS["B"],
            "color": CHART_COLORS["Category B"],
            "values": [category_matrix[issue_type]["B"] for issue_type in ISSUE_TYPES],
        },
        {
            "label": CATEGORY_LABELS["C"],
            "color": CHART_COLORS["Category C"],
            "values": [category_matrix[issue_type]["C"] for issue_type in ISSUE_TYPES],
        },
    ]

    political_group_map: dict[tuple[str, str], dict[str, Any]] = defaultdict(
        lambda: {
            "partyCombination": "",
            "subDivision": "",
            "issueCount": 0,
            "stations": set(),
            "categories": set(),
            "statuses": [],
        }
    )
    political_combo_totals: Counter[str] = Counter()

    for record in political_records:
        combo = record["partyCombination"]
        if not combo:
            continue

        subdivision = record["subDivision"] or infer_subdivision(record["policeStation"]) or "Unmapped"
        key = (combo, subdivision)
        group = political_group_map[key]
        group["partyCombination"] = combo
        group["subDivision"] = subdivision
        group["issueCount"] += 1
        group["stations"].add(record["policeStation"])
        if record["category"]:
            group["categories"].add(record["category"])
        if record["presentStatus"]:
            group["statuses"].append(record["presentStatus"])

        political_combo_totals[combo] += 1

    political_rows = []
    for group in political_group_map.values():
        categories = {category for category in CATEGORY_ORDER if category in group["categories"]}
        political_rows.append(
            {
                "partyCombination": group["partyCombination"],
                "subDivision": group["subDivision"],
                "policeStationsAffected": len(group["stations"]),
                "issueCount": group["issueCount"],
                "categories": "/".join(categories) if categories else "-",
                "presentStatus": political_status_label(categories, group["issueCount"], group["statuses"]),
            }
        )

    political_rows.sort(
        key=lambda row: (
            -row["issueCount"],
            -row["policeStationsAffected"],
            row["partyCombination"],
            row["subDivision"],
        )
    )
    political_rows = political_rows[:POLITICAL_BREAKDOWN_LIMIT]

    political_chart_rows = [
        {
            "partyCombination": combo,
            "issueCount": issue_count,
        }
        for combo, issue_count in sorted(
            political_combo_totals.items(),
            key=lambda item: (-item[1], item[0]),
        )[:POLITICAL_CHART_LIMIT]
    ]

    action_rows = []
    for position, record in enumerate(
        sorted(
            records,
            key=lambda row: (
                CATEGORY_SEVERITY_RANK.get(row["category"], 99),
                ISSUE_TYPE_PRIORITY.get(row["issueType"], 99),
                row["subDivision"],
                row["policeStation"],
                row["village"],
            ),
        )[:ACTION_TRACKER_LIMIT],
        start=1,
    ):
        action_rows.append(
            {
                "serial": position,
                "policeStation": record["policeStation"],
                "village": record["village"],
                "issueSummary": record["issueSummary"],
                "category": record["category"],
                "alertLevel": CATEGORY_ALERT_LABELS.get(record["category"], "MONITOR"),
                "actionTaken": record["actionTaken"],
                "presentStatus": record["presentStatus"],
            }
        )

    dashboard_payload = {
        "reportDate": report_date_obj.isoformat(),
        "reportDateLabel": report_date_label,
        "recordCount": len(records),
        "titles": {
            "executive": f"VILLAGE ISSUES & LAW AND ORDER - ELURU DISTRICT\nStatus Report as on {report_date_label} | Eluru District Police",
            "psWise": f"POLICE STATION WISE - ISSUE DETAILS\nEluru District | As on {report_date_label}",
            "category": "CATEGORY-WISE ISSUE ANALYSIS - ELURU DISTRICT\nDetailed breakdown by issue type and category (A/B/C)",
            "political": f"POLITICAL ISSUES - PARTY-WISE ANALYSIS\nEluru District | As on {report_date_label}",
            "action": f"VILLAGE ISSUE - ACTION STATUS TRACKER\nKey issues with present status | Eluru District | {report_date_label}",
        },
        "kpis": {
            "totalIssues": total_issues,
            "politicalIssues": political_total,
            "generalIssues": general_total,
            "casteCommunalIssues": caste_communal_total,
            "casteIssues": caste_total,
            "communalIssues": communal_total,
            "politicalInsight": executive_card_insight(
                political_total,
                total_issues,
                top_card_highest,
                tie_value=general_total,
            ),
            "generalInsight": executive_card_insight(
                general_total,
                total_issues,
                top_card_highest,
                tie_value=political_total,
            ),
            "casteCommunalInsight": f"Caste:{caste_total} | Communal:{communal_total}",
        },
        "executive": {
            "subDivisionRows": subdivision_rows,
            "subDivisionTotal": {
                "casteConflicts": caste_total,
                "politicalIssues": political_total,
                "communalIssues": communal_total,
                "generalIssues": general_total,
                "totalIssues": total_issues,
                "share": percentage_string(total_issues, total_issues, 0),
            },
            "issueDistribution": executive_distribution,
            "subDivisionChartSeries": executive_chart_series,
        },
        "psWise": {
            "rows": ps_rows,
            "total": {
                "casteConflicts": caste_total,
                "politicalIssues": political_total,
                "communalIssues": communal_total,
                "generalIssues": general_total,
                "totalIssues": total_issues,
            },
            "topStations": top_stations,
        },
        "categoryAnalysis": {
            "rows": category_rows,
            "totals": {
                "categoryA": category_totals["A"],
                "categoryB": category_totals["B"],
                "categoryC": category_totals["C"],
                "total": overall_category_total,
                "percent": percentage_string(overall_category_total, overall_category_total, 1),
            },
            "categoryChartSeries": category_chart_series,
            "categorySplit": category_split,
        },
        "politicalAnalysis": {
            "rows": political_rows,
            "chartRows": political_chart_rows,
        },
        "actionTracker": {
            "rows": action_rows,
        },
    }

    return dashboard_payload


def update_worksheet_range(
    worksheet: Any,
    start_row: int,
    end_row: int,
    values: list[list[Any]],
    start_col_index: int = 2,
) -> None:
    max_rows = end_row - start_row + 1
    for row_offset in range(max_rows):
        target_row = start_row + row_offset
        source_row = values[row_offset] if row_offset < len(values) else []
        for col_offset in range(len(source_row)):
            worksheet.cell(target_row, start_col_index + col_offset).value = source_row[col_offset]


def clear_cell_range(worksheet: Any, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            worksheet.cell(row, col).value = None


def build_dashboard_workbook(raw_records: list[dict[str, Any]] | None, report_date: Any = None) -> BytesIO:
    dashboard_data = build_dashboard_payload(raw_records, report_date)
    workbook = load_workbook(DASHBOARD_TEMPLATE_PATH)

    executive_sheet = workbook["Executive Dashboard"]
    executive_sheet["B2"] = dashboard_data["titles"]["executive"]
    executive_sheet["B8"] = dashboard_data["kpis"]["totalIssues"]
    executive_sheet["D8"] = dashboard_data["kpis"]["politicalIssues"]
    executive_sheet["F8"] = dashboard_data["kpis"]["generalIssues"]
    executive_sheet["H8"] = dashboard_data["kpis"]["casteCommunalIssues"]
    executive_sheet["B9"] = "Across all categories"
    executive_sheet["D9"] = dashboard_data["kpis"]["politicalInsight"]
    executive_sheet["F9"] = dashboard_data["kpis"]["generalInsight"]
    executive_sheet["H9"] = dashboard_data["kpis"]["casteCommunalInsight"]

    subdivision_rows = dashboard_data["executive"]["subDivisionRows"]
    for index, row in enumerate(subdivision_rows, start=13):
        executive_sheet[f"B{index}"] = row["name"]
        executive_sheet[f"C{index}"] = row["casteConflicts"]
        executive_sheet[f"E{index}"] = row["politicalIssues"]
        executive_sheet[f"F{index}"] = row["communalIssues"]
        executive_sheet[f"G{index}"] = row["generalIssues"]
        executive_sheet[f"H{index}"] = row["totalIssues"]
        executive_sheet[f"I{index}"] = row["share"]

    executive_totals = dashboard_data["executive"]["subDivisionTotal"]
    executive_sheet["B17"] = "GRAND TOTAL"
    executive_sheet["C17"] = executive_totals["casteConflicts"]
    executive_sheet["E17"] = executive_totals["politicalIssues"]
    executive_sheet["F17"] = executive_totals["communalIssues"]
    executive_sheet["G17"] = executive_totals["generalIssues"]
    executive_sheet["H17"] = executive_totals["totalIssues"]
    executive_sheet["I17"] = executive_totals["share"]

    for index, row in enumerate(subdivision_rows, start=39):
        executive_sheet[f"O{index}"] = row["shortName"]
        executive_sheet[f"P{index}"] = row["casteConflicts"]
        executive_sheet[f"Q{index}"] = row["politicalIssues"]
        executive_sheet[f"R{index}"] = row["communalIssues"]
        executive_sheet[f"S{index}"] = row["generalIssues"]
        executive_sheet[f"T{index}"] = row["totalIssues"]

    for index, row in enumerate(dashboard_data["executive"]["issueDistribution"], start=39):
        executive_sheet[f"V{index}"] = row["label"]
        executive_sheet[f"W{index}"] = row["value"]

    ps_sheet = workbook["PS-Wise Analysis"]
    ps_sheet["B1"] = dashboard_data["titles"]["psWise"]
    ps_rows = dashboard_data["psWise"]["rows"]
    for row_index, row in enumerate(ps_rows, start=6):
        ps_sheet[f"E{row_index}"] = row["casteConflicts"]
        ps_sheet[f"F{row_index}"] = row["politicalIssues"]
        ps_sheet[f"G{row_index}"] = row["communalIssues"]
        ps_sheet[f"H{row_index}"] = row["generalIssues"]
        ps_sheet[f"I{row_index}"] = row["totalIssues"]
        ps_sheet[f"J{row_index}"] = row["severity"]

    ps_totals = dashboard_data["psWise"]["total"]
    ps_sheet["E40"] = ps_totals["casteConflicts"]
    ps_sheet["F40"] = ps_totals["politicalIssues"]
    ps_sheet["G40"] = ps_totals["communalIssues"]
    ps_sheet["H40"] = ps_totals["generalIssues"]
    ps_sheet["I40"] = ps_totals["totalIssues"]

    clear_cell_range(ps_sheet, 81, 90, 13, 14)
    for row_index, row in enumerate(dashboard_data["psWise"]["topStations"], start=81):
        ps_sheet[f"M{row_index}"] = row["policeStation"]
        ps_sheet[f"N{row_index}"] = row["totalIssues"]

    category_sheet = workbook["Category Analysis"]
    category_sheet["B1"] = dashboard_data["titles"]["category"]
    for row_index, row in enumerate(dashboard_data["categoryAnalysis"]["rows"], start=11):
        category_sheet[f"B{row_index}"] = row["issueType"]
        category_sheet[f"C{row_index}"] = row["categoryA"]
        category_sheet[f"D{row_index}"] = row["categoryB"]
        category_sheet[f"E{row_index}"] = row["categoryC"]
        category_sheet[f"F{row_index}"] = row["total"]
        category_sheet[f"G{row_index}"] = row["percent"]

    category_totals = dashboard_data["categoryAnalysis"]["totals"]
    category_sheet["B15"] = "TOTAL"
    category_sheet["C15"] = category_totals["categoryA"]
    category_sheet["D15"] = category_totals["categoryB"]
    category_sheet["E15"] = category_totals["categoryC"]
    category_sheet["F15"] = category_totals["total"]
    category_sheet["G15"] = category_totals["percent"]

    for row_index, row in enumerate(dashboard_data["categoryAnalysis"]["rows"], start=81):
        category_sheet[f"J{row_index}"] = row["issueType"]
        category_sheet[f"K{row_index}"] = row["categoryA"]
        category_sheet[f"L{row_index}"] = row["categoryB"]
        category_sheet[f"M{row_index}"] = row["categoryC"]

    for row_index, row in enumerate(dashboard_data["categoryAnalysis"]["categorySplit"], start=87):
        category_sheet[f"J{row_index}"] = row["label"]
        category_sheet[f"K{row_index}"] = row["value"]

    political_sheet = workbook["Political Party Analysis"]
    political_sheet["B1"] = dashboard_data["titles"]["political"]
    clear_cell_range(political_sheet, 8, 17, 2, 7)
    for row_index, row in enumerate(dashboard_data["politicalAnalysis"]["rows"], start=8):
        political_sheet[f"B{row_index}"] = row["partyCombination"]
        political_sheet[f"C{row_index}"] = row["subDivision"]
        political_sheet[f"D{row_index}"] = row["policeStationsAffected"]
        political_sheet[f"E{row_index}"] = row["issueCount"]
        political_sheet[f"F{row_index}"] = row["categories"]
        political_sheet[f"G{row_index}"] = row["presentStatus"]

    clear_cell_range(political_sheet, 81, 87, 10, 11)
    for row_index, row in enumerate(dashboard_data["politicalAnalysis"]["chartRows"], start=81):
        political_sheet[f"J{row_index}"] = row["partyCombination"]
        political_sheet[f"K{row_index}"] = row["issueCount"]

    action_sheet = workbook["Action Status Tracker"]
    action_sheet["B1"] = dashboard_data["titles"]["action"]
    clear_cell_range(action_sheet, 6, 20, 2, 9)
    for row_index, row in enumerate(dashboard_data["actionTracker"]["rows"], start=6):
        action_sheet[f"B{row_index}"] = row["serial"]
        action_sheet[f"C{row_index}"] = row["policeStation"]
        action_sheet[f"D{row_index}"] = row["village"]
        action_sheet[f"E{row_index}"] = row["issueSummary"]
        action_sheet[f"F{row_index}"] = row["category"]
        action_sheet[f"G{row_index}"] = row["alertLevel"]
        action_sheet[f"H{row_index}"] = row["actionTaken"]
        action_sheet[f"I{row_index}"] = row["presentStatus"]

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
